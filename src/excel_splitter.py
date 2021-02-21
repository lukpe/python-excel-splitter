"""
A simple Tk app dividing Excel file basing on selected column values
"""
import tkinter as tk
from functools import partial
from pathlib import Path
from tkinter import filedialog, messagebox

import openpyxl as xl
from openpyxl import Workbook
from openpyxl.utils.exceptions import InvalidFileException


class App:
    """
    The main app class
    """

    def __init__(self):
        self.file_path = "None"
        self.input_file = {}
        padding_x = 10
        padding_y = 5
        separator_width = 350

        self.root = tk.Tk()
        self.root.winfo_toplevel().title("Excel file splitter")

        button_file = tk.Button(text="Choose file", command=self.choose_file, width=15)
        button_file.pack(padx=padding_x, pady=padding_y)

        self.label_file = tk.Label(text="File: " + self.file_path)
        self.label_file.pack(padx=padding_x, pady=padding_y, )

        separator_1 = tk.Frame(self.root, bg='black', height=1, width=separator_width)
        separator_1.pack(padx=padding_x, pady=padding_y)

        label_column = tk.Label(text="Column name:")
        label_column.pack(padx=padding_x, pady=padding_y)

        self.variable = tk.StringVar(self.root)
        self.variable.set("None")
        self.list_column = tk.OptionMenu(self.root, self.variable, "None")
        self.list_column.pack(padx=padding_x, pady=padding_y)

        separator_2 = tk.Frame(self.root, bg='black', height=1, width=separator_width)
        separator_2.pack(padx=padding_x, pady=padding_y)

        self.button_split = tk.Button(text="Split", command=self.split_workbook, state="disabled")
        self.button_split.pack(padx=padding_x, pady=padding_y)

        self.root.mainloop()

    def choose_file(self):
        """
        Choose file using native dialog window
        """
        try:
            self.file_path = filedialog.askopenfilename()
            self.label_file["text"] = "File: " + self.file_path
            self.input_file['workbook'] = xl.load_workbook(filename=self.file_path, data_only=True)
            self.input_file['worksheet'] = self.input_file['workbook'].worksheets[0]
            if self.update_list(self.input_file['worksheet']):
                self.button_split['state'] = "normal"
        except InvalidFileException:
            self.disable_split()
            self.label_file["text"] = "File: Incorrect"
            self.message_incorrect_file()

    def update_list(self, ws_in):
        """
        Show list of columns found in the input file (ws_in)
        """
        options = []
        empty_columns = True

        for col_in in range(1, ws_in.max_column + 1):
            column_name = ws_in.cell(column=col_in, row=1).value
            if column_name is not None:
                options.append(ws_in.cell(column=col_in, row=1).value)
                empty_columns = False
        if empty_columns:
            options.clear()
            options.append("None")
            self.disable_split()
            self.message_empty_file()
        menu = self.list_column["menu"]
        menu.delete(0, "end")
        for option in options:
            menu.add_command(label=option, command=partial(self.variable.set, option))
        self.variable.set(options[0])
        if not empty_columns:
            return True
        return False

    def split_workbook(self):
        """
        Split the workbook by the selected column values
        """
        try:
            ext_position = self.file_path.find(".")
            file_name = self.file_path[0:ext_position]
            file_ext = self.file_path[ext_position:]
            col_name = self.variable.get()
            sheet_name = self.input_file['worksheet'].title

            for row_in in range(2, self.input_file['worksheet'].max_row + 1):
                col_num = self.get_column_number(col_name)
                value = self.input_file['worksheet'].cell(column=col_num, row=row_in).value
                if value is None:
                    value = "EmptyValue"
                file_out = Path(file_name + "_" + value + file_ext)
                wb_out = xl.load_workbook(self.create_workbook(file_out))
                ws_out = wb_out.worksheets[0]
                ws_out.title = sheet_name
                row_out = ws_out.max_row + 1
                for col_out in range(1, self.input_file['worksheet'].max_column + 1):
                    ws_out.cell(column=col_out, row=row_out).value = \
                        self.input_file['worksheet'].cell(column=col_out, row=row_in).value
                wb_out.save(file_out)
                wb_out.close()
            self.input_file['workbook'].close()
            self.message_success()
            self.quit()
        except AttributeError:
            self.message_no_file()

    def get_column_number(self, col_name):
        """
        Return column number by getting its' name
        """
        for col_in in range(1, self.input_file['worksheet'].max_column + 1):
            if col_name in self.input_file['worksheet'].cell(column=col_in, row=1).value:
                return col_in
        return -1

    def create_workbook(self, file):
        """
        Create output workbook if it doesn't exist
        """
        if not file.exists():
            wb_out = Workbook()
            ws_out = wb_out.worksheets[0]
            for col_in in range(1, self.input_file['worksheet'].max_column + 1):
                ws_out.cell(column=col_in, row=1).value = \
                    self.input_file['worksheet'].cell(column=col_in, row=1).value
            wb_out.save(file)
        return file

    def disable_split(self):
        """
        Disable the 'Split' button
        """
        if self.button_split['state'] == "normal":
            self.button_split['state'] = "disabled"

    @staticmethod
    def message_success():
        """
        Success message
        """
        messagebox.showinfo(title="Success", message="The file has been split correctly")

    @staticmethod
    def message_no_file():
        """
        No file warning
        """
        messagebox.showwarning(title="Warning", message="Choose a correct file to split")

    @staticmethod
    def message_incorrect_file():
        """
        Wrong file format warning
        """
        messagebox.showwarning(title="Warning",
                               message="Choose a file in correct format (*.xls, *.xlsx)")

    @staticmethod
    def message_empty_file():
        """
        Empty file waring
        """
        messagebox.showwarning(title="Warning", message="Choose a file with at least one column")

    def quit(self):
        """
        Quit the app
        """
        self.root.destroy()


if __name__ == '__main__':
    app = App()
