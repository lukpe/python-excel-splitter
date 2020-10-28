import tkinter as tk
from pathlib import Path
from tkinter import filedialog, messagebox

import openpyxl as xl
from openpyxl import Workbook


class App:
    def __init__(self):
        self.root = tk.Tk()
        canvas1 = tk.Canvas(self.root, width=200, height=100)
        canvas1.pack()
        label1 = tk.Label(text="Column name:")
        canvas1.create_window(100, 20, window=label1)
        self.entry1 = tk.Entry(self.root)
        self.entry1.insert(0, "Test")
        canvas1.create_window(100, 40, window=self.entry1)
        button1 = tk.Button(text="Choose file", command=self.choose_file)
        canvas1.create_window(100, 75, window=button1)
        self.root.mainloop()

    def choose_file(self):
        file_path = filedialog.askopenfilename()
        self.split_workbook(file_path, self.entry1.get())
        self.message_success()
        self.quit()

    def split_workbook(self, file_path, col_name):
        ext_position = file_path.find(".")
        file_name = file_path[0:ext_position]
        file_ext = file_path[ext_position:]

        wb_in = xl.load_workbook(filename=file_name + file_ext, data_only=True)
        ws_in = wb_in.worksheets[0]

        for row_in in range(2, ws_in.max_row + 1):
            col_num = self.get_column_number(ws_in, col_name)
            value = ws_in.cell(column=col_num, row=row_in).value
            if value is None:
                value = "Empty"
            file_out = Path(file_name + "_" + value + file_ext)
            wb_out = xl.load_workbook(self.create_workbook(ws_in, file_out))
            ws_out = wb_out.worksheets[0]
            row_out = ws_out.max_row + 1
            for col_out in range(1, ws_in.max_column + 1):
                ws_out.cell(column=col_out, row=row_out).value = \
                    ws_in.cell(column=col_out, row=row_in).value
            wb_out.save(file_out)
            wb_out.close()
        wb_in.close()

    @staticmethod
    def get_column_number(ws_in, col_name):
        for col_in in range(1, ws_in.max_column + 1):
            if col_name in ws_in.cell(column=col_in, row=1).value:
                return col_in

    @staticmethod
    def create_workbook(ws_in, file):
        if not file.exists():
            wb_out = Workbook()
            ws_out = wb_out.worksheets[0]
            for col_in in range(1, ws_in.max_column + 1):
                ws_out.cell(column=col_in, row=1).value = ws_in.cell(column=col_in, row=1).value
            wb_out.save(file)
        return file

    @staticmethod
    def message_success():
        messagebox.showinfo(title="Success", message="File split correctly")

    def quit(self):
        self.root.destroy()


app = App()
